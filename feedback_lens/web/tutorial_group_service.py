from __future__ import annotations

import csv
import hashlib
import json
import re
import secrets
import sqlite3
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path

from werkzeug.security import generate_password_hash

from feedback_lens.web.common import record_audit_event
from feedback_lens.web.config import get_web_settings
from feedback_lens.web.errors import ApiError
from feedback_lens.web.mail import get_email_sender, mail_is_configured
from feedback_lens.web.security import can_administer_unit
from feedback_lens.web.storage import StoredUpload, remove_stored_upload


def _utc_text(value: datetime) -> str:
    return value.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def _unit_context(
    conn: sqlite3.Connection,
    actor_user_id: int,
    unit_offering_id: int,
) -> sqlite3.Row:
    row = conn.execute(
        """
        SELECT offering.unit_offering_id, course.organization_id,
               course.course_code, course.course_name
        FROM unit_offerings AS offering
        JOIN courses AS course ON course.course_id = offering.course_id
        WHERE offering.unit_offering_id = ?
        """,
        (unit_offering_id,),
    ).fetchone()
    if row is None:
        raise ApiError("unit_not_found", "Unit not found.", 404)
    if not can_administer_unit(conn, actor_user_id, unit_offering_id):
        raise ApiError(
            "unit_forbidden",
            "You are not authorised to manage this Unit.",
            403,
        )
    return row


def _group_context(
    conn: sqlite3.Connection,
    actor_user_id: int,
    tutorial_group_id: int,
) -> sqlite3.Row:
    row = conn.execute(
        """
        SELECT tutorial_group_id, unit_offering_id, group_code,
               group_name, active
        FROM tutorial_groups
        WHERE tutorial_group_id = ?
        """,
        (tutorial_group_id,),
    ).fetchone()
    if row is None:
        raise ApiError("tutorial_group_not_found", "Tutorial Group not found.", 404)
    _unit_context(conn, actor_user_id, int(row["unit_offering_id"]))
    return row


def create_tutorial_group(
    conn: sqlite3.Connection,
    actor_user_id: int,
    unit_offering_id: int,
    payload: dict,
) -> dict:
    _unit_context(conn, actor_user_id, unit_offering_id)
    group_code = str(payload.get("group_code") or "").strip()
    group_name = str(payload.get("group_name") or "").strip() or None
    if not group_code:
        raise ApiError(
            "tutorial_group_code_required",
            "Enter a Tutorial Group code.",
            422,
        )
    existing = conn.execute(
        """
        SELECT tutorial_group_id
        FROM tutorial_groups
        WHERE unit_offering_id = ? AND group_code = ?
        """,
        (unit_offering_id, group_code),
    ).fetchone()
    if existing is None:
        group_id = int(
            conn.execute(
                """
                INSERT INTO tutorial_groups
                    (unit_offering_id, group_code, group_name,
                     source, created_by_user_id)
                VALUES (?, ?, ?, 'manual', ?)
                """,
                (unit_offering_id, group_code, group_name, actor_user_id),
            ).lastrowid
        )
    else:
        group_id = int(existing["tutorial_group_id"])
        conn.execute(
            """
            UPDATE tutorial_groups
            SET group_name = COALESCE(?, group_name), active = 1,
                updated_at = CURRENT_TIMESTAMP
            WHERE tutorial_group_id = ?
            """,
            (group_name, group_id),
        )
    record_audit_event(
        conn,
        "tutorial_group.created",
        "tutorial_group",
        group_id,
        actor_user_id=actor_user_id,
        metadata={"unit_offering_id": unit_offering_id, "group_code": group_code},
    )
    conn.commit()
    return get_tutorial_group_overview(conn, actor_user_id, unit_offering_id)


def get_tutorial_group_overview(
    conn: sqlite3.Connection,
    actor_user_id: int,
    unit_offering_id: int,
) -> dict:
    _unit_context(conn, actor_user_id, unit_offering_id)
    group_rows = conn.execute(
        """
        SELECT
            group_row.tutorial_group_id,
            group_row.group_code,
            group_row.group_name,
            group_row.source,
            group_row.active,
            COUNT(DISTINCT membership.student_id) AS student_count
        FROM tutorial_groups AS group_row
        LEFT JOIN student_tutorial_memberships AS membership
          ON membership.tutorial_group_id = group_row.tutorial_group_id
         AND membership.active = 1
        WHERE group_row.unit_offering_id = ?
          AND group_row.active = 1
        GROUP BY group_row.tutorial_group_id
        ORDER BY lower(group_row.group_code), group_row.tutorial_group_id
        """,
        (unit_offering_id,),
    ).fetchall()
    staff_rows = conn.execute(
        """
        SELECT link.tutorial_group_id, link.user_id,
               user.email, user.display_name, user.account_status,
               link.allocation_weight
        FROM tutorial_group_staff AS link
        JOIN tutorial_groups AS group_row
          ON group_row.tutorial_group_id = link.tutorial_group_id
        JOIN users AS user ON user.user_id = link.user_id
        WHERE group_row.unit_offering_id = ?
          AND group_row.active = 1
          AND link.active = 1
        ORDER BY lower(user.email), user.user_id
        """,
        (unit_offering_id,),
    ).fetchall()
    staff_by_group: dict[int, list[dict]] = {}
    for row in staff_rows:
        staff_by_group.setdefault(int(row["tutorial_group_id"]), []).append(dict(row))

    groups = []
    for row in group_rows:
        item = dict(row)
        staff = staff_by_group.get(int(row["tutorial_group_id"]), [])
        active_staff = [person for person in staff if person["account_status"] == "active"]
        pending_staff = [person for person in staff if person["account_status"] == "pending"]
        item.update(
            {
                "staff": staff,
                "active_staff_count": len(active_staff),
                "pending_staff_count": len(pending_staff),
                "ready": bool(active_staff) and not pending_staff,
            }
        )
        groups.append(item)

    ungrouped_count = int(
        conn.execute(
            """
            SELECT COUNT(*)
            FROM student_enrolments AS enrolment
            WHERE enrolment.unit_offering_id = ?
              AND enrolment.status = 'active'
              AND NOT EXISTS (
                  SELECT 1
                  FROM student_tutorial_memberships AS membership
                  WHERE membership.unit_offering_id = enrolment.unit_offering_id
                    AND membership.student_id = enrolment.student_id
                    AND membership.active = 1
              )
            """,
            (unit_offering_id,),
        ).fetchone()[0]
    )
    unit_staff = [
        dict(row)
        for row in conn.execute(
            """
            SELECT DISTINCT user.user_id, user.email, user.display_name,
                   user.account_status
            FROM unit_role_assignments AS role
            JOIN users AS user ON user.user_id = role.user_id
            WHERE role.unit_offering_id = ?
              AND role.role = 'staff'
              AND role.active = 1
              AND user.account_status = 'active'
            ORDER BY lower(user.email), user.user_id
            """,
            (unit_offering_id,),
        )
    ]
    imports = [
        dict(row)
        for row in conn.execute(
            """
            SELECT tutorial_group_import_id, source_file_name, status,
                   total_row_count, valid_row_count, invalid_row_count,
                   assigned_count, moved_count, unchanged_count,
                   previewed_at, applied_at
            FROM tutorial_group_imports
            WHERE unit_offering_id = ?
            ORDER BY tutorial_group_import_id DESC
            LIMIT 5
            """,
            (unit_offering_id,),
        )
    ]
    return {
        "groups": groups,
        "unit_staff": unit_staff,
        "ungrouped_student_count": ungrouped_count,
        "ready_group_count": sum(group["ready"] for group in groups),
        "incomplete_group_count": sum(not group["ready"] for group in groups),
        "recent_imports": imports,
    }


def add_tutorial_group_staff(
    conn: sqlite3.Connection,
    actor_user_id: int,
    tutorial_group_id: int,
    staff_user_id: int,
) -> dict:
    group = _group_context(conn, actor_user_id, tutorial_group_id)
    eligible = conn.execute(
        """
        SELECT user.user_id, user.account_status
        FROM users AS user
        WHERE user.user_id = ?
          AND user.role IN ('admin', 'lead_lecturer', 'educator')
          AND user.account_status = 'active'
          AND EXISTS (
              SELECT 1
              FROM unit_role_assignments AS role
              WHERE role.unit_offering_id = ?
                AND role.user_id = user.user_id
                AND role.role IN ('unit_admin', 'staff')
                AND role.active = 1
          )
        """,
        (staff_user_id, group["unit_offering_id"]),
    ).fetchone()
    if eligible is None:
        raise ApiError(
            "tutorial_staff_invalid",
            "Choose an active Staff member in this Unit.",
            422,
        )
    conn.execute(
        """
        INSERT INTO tutorial_group_staff
            (tutorial_group_id, user_id, assigned_by_user_id)
        VALUES (?, ?, ?)
        ON CONFLICT(tutorial_group_id, user_id)
        DO UPDATE SET active = 1, ended_at = NULL,
                      assigned_at = CURRENT_TIMESTAMP,
                      assigned_by_user_id = excluded.assigned_by_user_id
        """,
        (tutorial_group_id, staff_user_id, actor_user_id),
    )
    record_audit_event(
        conn,
        "tutorial_group.staff_added",
        "tutorial_group",
        tutorial_group_id,
        actor_user_id=actor_user_id,
        metadata={"staff_user_id": staff_user_id},
    )
    conn.commit()
    return get_tutorial_group_overview(
        conn,
        actor_user_id,
        int(group["unit_offering_id"]),
    )


def remove_tutorial_group_staff(
    conn: sqlite3.Connection,
    actor_user_id: int,
    tutorial_group_id: int,
    staff_user_id: int,
) -> dict:
    group = _group_context(conn, actor_user_id, tutorial_group_id)
    conn.execute(
        """
        UPDATE tutorial_group_staff
        SET active = 0, ended_at = CURRENT_TIMESTAMP
        WHERE tutorial_group_id = ? AND user_id = ? AND active = 1
        """,
        (tutorial_group_id, staff_user_id),
    )
    record_audit_event(
        conn,
        "tutorial_group.staff_removed",
        "tutorial_group",
        tutorial_group_id,
        actor_user_id=actor_user_id,
        metadata={"staff_user_id": staff_user_id},
    )
    conn.commit()
    return get_tutorial_group_overview(
        conn,
        actor_user_id,
        int(group["unit_offering_id"]),
    )


def set_tutorial_staff_groups(
    conn: sqlite3.Connection,
    actor_user_id: int,
    unit_offering_id: int,
    staff_user_id: int,
    tutorial_group_ids: list[int],
) -> dict:
    _unit_context(conn, actor_user_id, unit_offering_id)
    eligible = conn.execute(
        """
        SELECT user.user_id
        FROM users AS user
        WHERE user.user_id = ?
          AND user.role IN ('admin', 'lead_lecturer', 'educator')
          AND user.account_status = 'active'
          AND EXISTS (
              SELECT 1
              FROM unit_role_assignments AS role
              WHERE role.unit_offering_id = ?
                AND role.user_id = user.user_id
                AND role.role IN ('unit_admin', 'staff')
                AND role.active = 1
          )
        """,
        (staff_user_id, unit_offering_id),
    ).fetchone()
    if eligible is None:
        raise ApiError(
            "tutorial_staff_invalid",
            "Choose an active Staff member in this Unit.",
            422,
        )
    if not isinstance(tutorial_group_ids, list):
        raise ApiError(
            "tutorial_group_invalid",
            "Tutorial Groups must be supplied as a list.",
            422,
        )
    try:
        requested_ids = sorted({int(value) for value in tutorial_group_ids})
    except (TypeError, ValueError) as error:
        raise ApiError(
            "tutorial_group_invalid",
            "One or more Tutorial Groups are invalid.",
            422,
        ) from error
    available_ids = {
        int(row["tutorial_group_id"])
        for row in conn.execute(
            """
            SELECT tutorial_group_id
            FROM tutorial_groups
            WHERE unit_offering_id = ? AND active = 1
            """,
            (unit_offering_id,),
        )
    }
    if not set(requested_ids).issubset(available_ids):
        raise ApiError(
            "tutorial_group_invalid",
            "One or more Tutorial Groups are no longer available.",
            409,
        )
    current_ids = {
        int(row["tutorial_group_id"])
        for row in conn.execute(
            """
            SELECT link.tutorial_group_id
            FROM tutorial_group_staff AS link
            JOIN tutorial_groups AS group_row
              ON group_row.tutorial_group_id = link.tutorial_group_id
            WHERE group_row.unit_offering_id = ?
              AND link.user_id = ?
              AND link.active = 1
            """,
            (unit_offering_id, staff_user_id),
        )
    }
    requested_set = set(requested_ids)
    conn.execute("BEGIN")
    try:
        for group_id in sorted(current_ids - requested_set):
            conn.execute(
                """
                UPDATE tutorial_group_staff
                SET active = 0, ended_at = CURRENT_TIMESTAMP
                WHERE tutorial_group_id = ? AND user_id = ? AND active = 1
                """,
                (group_id, staff_user_id),
            )
        for group_id in requested_ids:
            conn.execute(
                """
                INSERT INTO tutorial_group_staff
                    (tutorial_group_id, user_id, assigned_by_user_id)
                VALUES (?, ?, ?)
                ON CONFLICT(tutorial_group_id, user_id)
                DO UPDATE SET active = 1, ended_at = NULL,
                              assigned_at = CURRENT_TIMESTAMP,
                              assigned_by_user_id = excluded.assigned_by_user_id
                """,
                (group_id, staff_user_id, actor_user_id),
            )
        record_audit_event(
            conn,
            "tutorial_group.staff_groups_updated",
            "user",
            staff_user_id,
            actor_user_id=actor_user_id,
            metadata={
                "unit_offering_id": unit_offering_id,
                "tutorial_group_ids": requested_ids,
                "added_group_ids": sorted(requested_set - current_ids),
                "removed_group_ids": sorted(current_ids - requested_set),
            },
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return get_tutorial_group_overview(
        conn,
        actor_user_id,
        unit_offering_id,
    )


def invite_tutorial_group_staff(
    conn: sqlite3.Connection,
    actor_user_id: int,
    unit_offering_id: int,
    payload: dict,
) -> dict:
    _unit_context(conn, actor_user_id, unit_offering_id)
    email = str(payload.get("email") or "").strip().casefold()
    display_name = str(payload.get("display_name") or "").strip()
    raw_group_ids = payload.get("tutorial_group_ids")
    if not email or "@" not in email:
        raise ApiError("staff_email_invalid", "Enter a valid Staff email.", 422)
    if not isinstance(raw_group_ids, list) or not raw_group_ids:
        raise ApiError(
            "tutorial_group_required",
            "Choose at least one Tutorial Group.",
            422,
        )
    try:
        group_ids = sorted({int(value) for value in raw_group_ids})
    except (TypeError, ValueError) as error:
        raise ApiError(
            "tutorial_group_invalid",
            "One or more Tutorial Groups are invalid.",
            422,
        ) from error
    found_groups = {
        int(row["tutorial_group_id"])
        for row in conn.execute(
            f"""
            SELECT tutorial_group_id
            FROM tutorial_groups
            WHERE unit_offering_id = ? AND active = 1
              AND tutorial_group_id IN ({', '.join('?' for _ in group_ids)})
            """,
            (unit_offering_id, *group_ids),
        )
    }
    if found_groups != set(group_ids):
        raise ApiError(
            "tutorial_group_invalid",
            "One or more Tutorial Groups are no longer available.",
            409,
        )

    user = conn.execute(
        "SELECT * FROM users WHERE lower(email) = lower(?)",
        (email,),
    ).fetchone()
    created = user is None
    if user is not None and user["role"] == "student":
        raise ApiError(
            "staff_email_conflict",
            "This email belongs to a student account.",
            409,
        )
    if user is not None and user["account_status"] == "disabled":
        raise ApiError(
            "staff_account_disabled",
            "This Staff account is disabled.",
            409,
        )
    if user is None:
        user_id = int(
            conn.execute(
                """
                INSERT INTO users
                    (email, password_hash, role, display_name, account_status)
                VALUES (?, ?, 'educator', ?, 'pending')
                """,
                (
                    email,
                    generate_password_hash(secrets.token_urlsafe(32)),
                    display_name or None,
                ),
            ).lastrowid
        )
        account_status = "pending"
    else:
        user_id = int(user["user_id"])
        account_status = str(user["account_status"])
        if display_name and not user["display_name"]:
            conn.execute(
                "UPDATE users SET display_name = ? WHERE user_id = ?",
                (display_name, user_id),
            )
    conn.execute(
        """
        INSERT INTO unit_role_assignments
            (unit_offering_id, user_id, role, assigned_by_user_id)
        VALUES (?, ?, 'staff', ?)
        ON CONFLICT(unit_offering_id, user_id, role)
        DO UPDATE SET active = 1, ended_at = NULL,
                      assigned_at = CURRENT_TIMESTAMP,
                      assigned_by_user_id = excluded.assigned_by_user_id
        """,
        (unit_offering_id, user_id, actor_user_id),
    )
    for group_id in group_ids:
        conn.execute(
            """
            INSERT INTO tutorial_group_staff
                (tutorial_group_id, user_id, assigned_by_user_id)
            VALUES (?, ?, ?)
            ON CONFLICT(tutorial_group_id, user_id)
            DO UPDATE SET active = 1, ended_at = NULL,
                          assigned_at = CURRENT_TIMESTAMP,
                          assigned_by_user_id = excluded.assigned_by_user_id
            """,
            (group_id, user_id, actor_user_id),
        )

    activation_url = None
    email_sent = False
    if account_status == "pending":
        conn.execute(
            """
            UPDATE staff_activation_invites
            SET status = 'revoked'
            WHERE user_id = ? AND status = 'pending'
            """,
            (user_id,),
        )
        token = secrets.token_urlsafe(32)
        expires_at = _utc_text(
            datetime.now(timezone.utc)
            + timedelta(hours=get_web_settings().activation_ttl_hours)
        )
        invite_id = int(
            conn.execute(
                """
                INSERT INTO staff_activation_invites
                    (user_id, unit_offering_id, token_hash,
                     issued_by_user_id, expires_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    user_id,
                    unit_offering_id,
                    hashlib.sha256(token.encode("utf-8")).hexdigest(),
                    actor_user_id,
                    expires_at,
                ),
            ).lastrowid
        )
        activation_url = (
            f"{get_web_settings().public_base_url}/account/staff-activate/{token}"
        )
        record_audit_event(
            conn,
            "staff.invited",
            "staff_activation_invite",
            invite_id,
            actor_user_id=actor_user_id,
            metadata={
                "staff_user_id": user_id,
                "unit_offering_id": unit_offering_id,
                "tutorial_group_ids": group_ids,
            },
        )
    else:
        record_audit_event(
            conn,
            "tutorial_group.staff_added",
            "user",
            user_id,
            actor_user_id=actor_user_id,
            metadata={
                "unit_offering_id": unit_offering_id,
                "tutorial_group_ids": group_ids,
            },
        )
    conn.commit()

    if activation_url and mail_is_configured():
        try:
            get_email_sender().send(
                email,
                "Activate your Feedback Lens Staff account",
                (
                    f"Hello {display_name or 'Staff member'},\n\n"
                    "You have been invited to provide feedback in Feedback Lens.\n\n"
                    f"Activate your account here:\n\n{activation_url}\n\n"
                    "If you were not expecting this invitation, contact your Unit Admin."
                ),
            )
            email_sent = True
        except Exception:
            email_sent = False
    return {
        "status": "invited" if account_status == "pending" else "linked",
        "created": created,
        "user_id": user_id,
        "email_sent": email_sent,
        "activation_url": activation_url,
        "tutorial_groups": get_tutorial_group_overview(
            conn,
            actor_user_id,
            unit_offering_id,
        ),
    }


def staff_activation_is_valid(conn: sqlite3.Connection, token: str) -> bool:
    token_hash = hashlib.sha256(token.encode("utf-8")).hexdigest()
    return (
        conn.execute(
            """
            SELECT 1
            FROM staff_activation_invites AS invite
            JOIN users AS user ON user.user_id = invite.user_id
            WHERE invite.token_hash = ?
              AND invite.status = 'pending'
              AND invite.expires_at > CURRENT_TIMESTAMP
              AND user.account_status = 'pending'
            """,
            (token_hash,),
        ).fetchone()
        is not None
    )


def complete_staff_activation(
    conn: sqlite3.Connection,
    token: str,
    password: str,
) -> int:
    if len(password) < 12:
        raise ApiError(
            "password_too_short",
            "Password must contain at least 12 characters.",
            422,
        )
    token_hash = hashlib.sha256(token.encode("utf-8")).hexdigest()
    invite = conn.execute(
        """
        SELECT invite.*, user.account_status
        FROM staff_activation_invites AS invite
        JOIN users AS user ON user.user_id = invite.user_id
        WHERE invite.token_hash = ?
          AND invite.status = 'pending'
          AND invite.expires_at > CURRENT_TIMESTAMP
        """,
        (token_hash,),
    ).fetchone()
    if invite is None or invite["account_status"] != "pending":
        raise ApiError(
            "token_invalid",
            "This Staff activation link is invalid or has expired.",
            409,
        )
    user_id = int(invite["user_id"])
    conn.execute(
        """
        UPDATE users
        SET password_hash = ?, account_status = 'active',
            session_version = session_version + 1
        WHERE user_id = ?
        """,
        (generate_password_hash(password), user_id),
    )
    conn.execute(
        """
        UPDATE staff_activation_invites
        SET status = 'accepted', consumed_at = CURRENT_TIMESTAMP
        WHERE staff_activation_invite_id = ?
        """,
        (invite["staff_activation_invite_id"],),
    )
    record_audit_event(
        conn,
        "staff.account_activated",
        "user",
        user_id,
        actor_user_id=user_id,
        metadata={"unit_offering_id": invite["unit_offering_id"]},
    )
    conn.commit()
    return user_id


def _header_key(value: str) -> str:
    return "_".join(
        part for part in "".join(
            character.lower() if character.isalnum() else " "
            for character in value.strip()
        ).split() if part
    )


def _read_group_csv(path: str | Path) -> tuple[list[str], list[dict]]:
    try:
        with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = [str(value) for value in (reader.fieldnames or [])]
            rows = [dict(row) for row in reader]
    except UnicodeDecodeError as error:
        raise ApiError(
            "tutorial_csv_encoding",
            "The Tutorial Group CSV must use UTF-8 encoding.",
            422,
        ) from error
    if not headers:
        raise ApiError(
            "tutorial_csv_empty",
            "The Tutorial Group CSV has no header row.",
            422,
        )
    return headers, rows


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_excel_sheets(
    path: str | Path,
    extension: str,
) -> list[tuple[str, list[list]]]:
    try:
        if extension == ".xls":
            import xlrd

            workbook = xlrd.open_workbook(str(path), on_demand=True)
            try:
                return [
                    (
                        sheet.name,
                        [
                            [sheet.cell_value(row, column) for column in range(sheet.ncols)]
                            for row in range(sheet.nrows)
                        ],
                    )
                    for sheet in workbook.sheets()
                ]
            finally:
                workbook.release_resources()
        if extension == ".xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
            )
            try:
                return [
                    (
                        sheet.title,
                        [list(row) for row in sheet.iter_rows(values_only=True)],
                    )
                    for sheet in workbook.worksheets
                ]
            finally:
                workbook.close()
    except ImportError as error:
        raise ApiError(
            "tutorial_excel_dependency",
            "Excel import requires the xlrd and openpyxl packages.",
            503,
        ) from error
    except Exception as error:
        raise ApiError(
            "tutorial_excel_invalid",
            "The Allocate+ workbook could not be read. Upload a valid .xls or .xlsx file.",
            422,
        ) from error
    raise ApiError(
        "tutorial_excel_type",
        "Allocate+ membership files must use .xls or .xlsx format.",
        422,
    )


def _worksheet_group_code(sheet_name: str, rows: list[list]) -> str:
    if len(rows) > 2 and rows[2]:
        value = _cell_text(rows[2][0])
        if value:
            return value
    return re.sub(r"\(\d+\)$", "", sheet_name).strip()


def _activity_key(group_code: str) -> str:
    return re.split(r"[-_\s]", group_code, maxsplit=1)[0].casefold()


def _read_allocate_workbook(
    path: str | Path,
    extension: str,
    activity_type: str,
) -> tuple[list[str], list[dict]]:
    selected_activity = str(activity_type or "Applied").strip()
    if not selected_activity or len(selected_activity) > 80:
        raise ApiError(
            "tutorial_activity_invalid",
            "Enter a valid Allocate+ activity type.",
            422,
        )
    source_rows: list[dict] = []
    discovered_groups: list[str] = []
    selected_groups: list[str] = []
    source_number = 2
    student_keys = {
        "student_code",
        "student_id",
        "student_identifier",
        "student_number",
        "username",
    }
    for sheet_name, rows in _read_excel_sheets(path, extension):
        group_code = _worksheet_group_code(sheet_name, rows)
        if group_code:
            discovered_groups.append(group_code)
        if not group_code or _activity_key(group_code) != selected_activity.casefold():
            continue
        selected_groups.append(group_code)
        header_index = None
        student_column_index = None
        for index, row in enumerate(rows[:50]):
            keys = [_header_key(_cell_text(value)) for value in row]
            match = next(
                (position for position, key in enumerate(keys) if key in student_keys),
                None,
            )
            if match is not None:
                header_index = index
                student_column_index = match
                break
        if header_index is None or student_column_index is None:
            raise ApiError(
                "tutorial_excel_columns",
                f"Worksheet '{sheet_name}' has no student_code column.",
                422,
            )
        schedule = (
            _cell_text(rows[2][1])
            if len(rows) > 2 and len(rows[2]) > 1
            else ""
        )
        location = ""
        for metadata_row in rows[: min(header_index, 10)]:
            for value in metadata_row:
                text = _cell_text(value)
                if text.casefold().startswith("location:"):
                    location = text.split(":", 1)[1].strip()
                    break
            if location:
                break
        group_name = " · ".join(value for value in (schedule, location) if value)
        headers = []
        seen_headers: Counter[str] = Counter()
        for column, value in enumerate(rows[header_index]):
            label = _cell_text(value) or f"column_{column + 1}"
            seen_headers[label] += 1
            if seen_headers[label] > 1:
                label = f"{label}_{seen_headers[label]}"
            headers.append(label)
        for sheet_row_number, values in enumerate(
            rows[header_index + 1 :],
            start=header_index + 2,
        ):
            student_identifier = (
                _cell_text(values[student_column_index])
                if student_column_index < len(values)
                else ""
            )
            if not student_identifier:
                continue
            raw = {
                header: _cell_text(values[column]) if column < len(values) else ""
                for column, header in enumerate(headers)
            }
            raw.update(
                {
                    "tutorial_group": group_code,
                    "_source_sheet": sheet_name,
                    "_source_sheet_row": sheet_row_number,
                    "_source_row_number": source_number,
                    "_activity_type": selected_activity,
                    "_group_name": group_name or group_code,
                }
            )
            source_rows.append(raw)
            source_number += 1
    if not selected_groups:
        available = ", ".join(discovered_groups[:12]) or "none"
        raise ApiError(
            "tutorial_activity_missing",
            f"No '{selected_activity}' worksheets were found. Available classes: {available}.",
            422,
        )
    if not source_rows:
        raise ApiError(
            "tutorial_excel_empty",
            f"The '{selected_activity}' worksheets contain no student records.",
            422,
        )
    headers = list(source_rows[0])
    return headers, source_rows


def _read_group_source(
    upload: StoredUpload,
    activity_type: str,
) -> tuple[list[str], list[dict]]:
    if upload.extension == ".csv":
        return _read_group_csv(upload.storage_path)
    return _read_allocate_workbook(
        upload.storage_path,
        upload.extension,
        activity_type,
    )


def _first_source_column(
    by_key: dict[str, str],
    candidates: tuple[str, ...],
) -> str | None:
    return next((by_key[key] for key in candidates if key in by_key), None)


def _embedded_registration_columns(
    by_key: dict[str, str],
) -> dict[str, str] | None:
    columns = {
        "last_name": _first_source_column(
            by_key,
            ("last_name", "surname", "family_name"),
        ),
        "preferred_name": _first_source_column(
            by_key,
            ("preferred_name", "first_name", "given_name"),
        ),
        "email": _first_source_column(
            by_key,
            ("email_address", "institution_email", "institutional_email", "email"),
        ),
    }
    if any(value is None for value in columns.values()):
        return None
    return {key: str(value) for key, value in columns.items()}


def _registration_values(raw: dict, columns: dict[str, str] | None) -> dict:
    if columns is None:
        return {
            "available": False,
            "full_name": None,
            "email": None,
            "student_action": "existing",
            "enrolment_action": "unchanged",
        }
    preferred_name = str(raw.get(columns["preferred_name"]) or "").strip()
    last_name = str(raw.get(columns["last_name"]) or "").strip()
    return {
        "available": True,
        "full_name": " ".join(
            value for value in (preferred_name, last_name) if value
        ),
        "email": str(raw.get(columns["email"]) or "").strip().casefold(),
        "student_action": "pending",
        "enrolment_action": "pending",
    }


def _stored_registration(raw_data_json: str | None) -> dict:
    try:
        raw = json.loads(raw_data_json or "{}")
    except json.JSONDecodeError:
        return {"available": False}
    registration = raw.get("_registration")
    return registration if isinstance(registration, dict) else {"available": False}


def _apply_embedded_student_registration(
    conn: sqlite3.Connection,
    unit_offering_id: int,
    tutorial_group_import_id: int,
    identifier: str,
    registration: dict,
) -> int:
    full_name = str(registration.get("full_name") or "").strip()
    email = str(registration.get("email") or "").strip().casefold()
    if not identifier or not full_name or not email or "@" not in email:
        raise ApiError(
            "tutorial_student_identity_invalid",
            "The Applied import contains incomplete student identity data.",
            409,
        )
    student = conn.execute(
        """
        SELECT * FROM students
        WHERE lower(institution_student_identifier) = lower(?)
        """,
        (identifier,),
    ).fetchone()
    email_owner = conn.execute(
        """
        SELECT student_id FROM students
        WHERE lower(institution_email) = lower(?)
        """,
        (email,),
    ).fetchone()
    if email_owner is not None and (
        student is None or int(email_owner["student_id"]) != int(student["student_id"])
    ):
        raise ApiError(
            "tutorial_student_email_conflict",
            f"Student email {email} now belongs to another student.",
            409,
        )
    user_owner = conn.execute(
        "SELECT user_id FROM users WHERE lower(email) = lower(?)",
        (email,),
    ).fetchone()
    linked_user_id = int(student["user_id"]) if student and student["user_id"] else None
    if user_owner is not None and int(user_owner["user_id"]) != linked_user_id:
        raise ApiError(
            "tutorial_student_email_conflict",
            f"Student email {email} now belongs to another account.",
            409,
        )
    if student is None:
        student_id = int(
            conn.execute(
                """
                INSERT INTO students
                    (institution_student_identifier, full_name, institution_email)
                VALUES (?, ?, ?)
                """,
                (identifier, full_name, email),
            ).lastrowid
        )
    else:
        student_id = int(student["student_id"])
        if linked_user_id is not None:
            conn.execute(
                """
                UPDATE users
                SET email = ?, display_name = ?, updated_at = CURRENT_TIMESTAMP
                WHERE user_id = ?
                """,
                (email, full_name, linked_user_id),
            )
        conn.execute(
            """
            UPDATE students
            SET full_name = ?, institution_email = ?, updated_at = CURRENT_TIMESTAMP
            WHERE student_id = ?
            """,
            (full_name, email, student_id),
        )
    conn.execute(
        """
        INSERT INTO student_enrolments
            (unit_offering_id, student_id, status, source, source_reference)
        VALUES (?, ?, 'active', 'moodle', ?)
        ON CONFLICT(unit_offering_id, student_id)
        DO UPDATE SET status = 'active', source = 'moodle',
                      source_reference = excluded.source_reference,
                      ended_at = NULL
        """,
        (
            unit_offering_id,
            student_id,
            f"tutorial_group_import:{tutorial_group_import_id}",
        ),
    )
    return student_id


def import_tutorial_group_staff(
    conn: sqlite3.Connection,
    actor_user_id: int,
    unit_offering_id: int,
    upload: StoredUpload,
) -> dict:
    _unit_context(conn, actor_user_id, unit_offering_id)
    headers, raw_rows = _read_group_csv(upload.storage_path)
    by_key = {_header_key(header): header for header in headers}
    group_column = next(
        (
            by_key[key]
            for key in ("tutorial_group", "tutorial", "group", "class")
            if key in by_key
        ),
        None,
    )
    staff_column = next(
        (
            by_key[key]
            for key in ("staff_email", "tutor_email", "email")
            if key in by_key
        ),
        None,
    )
    if group_column is None or staff_column is None:
        raise ApiError(
            "tutorial_staff_csv_columns",
            "Staff mapping CSV columns must include tutorial_group and staff_email.",
            422,
        )
    groups = {
        str(row["group_code"]).casefold(): row
        for row in conn.execute(
            """
            SELECT tutorial_group_id, group_code
            FROM tutorial_groups
            WHERE unit_offering_id = ? AND active = 1
            """,
            (unit_offering_id,),
        )
    }
    staff = {
        str(row["email"]).casefold(): row
        for row in conn.execute(
            """
            SELECT user.user_id, user.email, user.account_status
            FROM users AS user
            WHERE user.role IN ('admin', 'lead_lecturer', 'educator')
              AND user.account_status = 'active'
            """,
        )
    }
    prepared = []
    relationship_counts: Counter[tuple[str, str]] = Counter()
    for number, raw in enumerate(raw_rows, start=2):
        group_code = str(raw.get(group_column) or "").strip()
        staff_email = str(raw.get(staff_column) or "").strip().casefold()
        if not group_code and not staff_email:
            continue
        key = (group_code.casefold(), staff_email)
        relationship_counts[key] += 1
        prepared.append(
            {
                "source_row_number": number,
                "group_code": group_code,
                "staff_email": staff_email,
                "raw": raw,
                "key": key,
            }
        )
    if not prepared:
        raise ApiError(
            "tutorial_staff_csv_empty",
            "The Staff mapping CSV contains no mappings.",
            422,
        )
    invalid_rows = []
    valid_rows = []
    for row in prepared:
        error = None
        group = groups.get(row["group_code"].casefold())
        person = staff.get(row["staff_email"])
        if not row["group_code"]:
            error = "Tutorial Group is required."
        elif group is None:
            error = "Tutorial Group does not exist in this Unit."
        elif not row["staff_email"] or "@" not in row["staff_email"]:
            error = "A valid Staff email is required."
        elif person is None:
            error = "Staff email must belong to an active Educator account."
        elif relationship_counts[row["key"]] > 1:
            error = "This Staff-to-Group mapping is duplicated in the CSV."
        if error:
            invalid_rows.append(
                {
                    "source_row_number": row["source_row_number"],
                    "tutorial_group": row["group_code"],
                    "staff_email": row["staff_email"],
                    "validation_error": error,
                }
            )
        else:
            valid_rows.append(
                {
                    **row,
                    "tutorial_group_id": int(group["tutorial_group_id"]),
                    "staff_user_id": int(person["user_id"]),
                }
            )
    if invalid_rows:
        raise ApiError(
            "tutorial_staff_csv_invalid",
            "Fix every invalid Staff mapping before importing the file.",
            422,
            {
                "invalid_row_count": len(invalid_rows),
                "rows": invalid_rows[:100],
            },
        )
    existing = {
        (int(row["tutorial_group_id"]), int(row["user_id"]))
        for row in conn.execute(
            """
            SELECT link.tutorial_group_id, link.user_id
            FROM tutorial_group_staff AS link
            JOIN tutorial_groups AS group_row
              ON group_row.tutorial_group_id = link.tutorial_group_id
            WHERE group_row.unit_offering_id = ? AND link.active = 1
            """,
            (unit_offering_id,),
        )
    }
    added_count = sum(
        (row["tutorial_group_id"], row["staff_user_id"]) not in existing
        for row in valid_rows
    )
    existing_unit_staff_ids = {
        int(row["user_id"])
        for row in conn.execute(
            """
            SELECT user_id
            FROM unit_role_assignments
            WHERE unit_offering_id = ?
              AND role IN ('unit_admin', 'staff')
              AND active = 1
            """,
            (unit_offering_id,),
        )
    }
    imported_staff_ids = {int(row["staff_user_id"]) for row in valid_rows}
    unit_staff_ids_to_add = imported_staff_ids - existing_unit_staff_ids
    conn.execute("BEGIN")
    try:
        conn.executemany(
            """
            INSERT INTO unit_role_assignments
                (unit_offering_id, user_id, role, assigned_by_user_id)
            VALUES (?, ?, 'staff', ?)
            ON CONFLICT(unit_offering_id, user_id, role)
            DO UPDATE SET active = 1, ended_at = NULL,
                          assigned_at = CURRENT_TIMESTAMP,
                          assigned_by_user_id = excluded.assigned_by_user_id
            """,
            [
                (unit_offering_id, staff_user_id, actor_user_id)
                for staff_user_id in sorted(unit_staff_ids_to_add)
            ],
        )
        conn.executemany(
            """
            INSERT INTO tutorial_group_staff
                (tutorial_group_id, user_id, assigned_by_user_id)
            VALUES (?, ?, ?)
            ON CONFLICT(tutorial_group_id, user_id)
            DO UPDATE SET active = 1, ended_at = NULL,
                          assigned_at = CURRENT_TIMESTAMP,
                          assigned_by_user_id = excluded.assigned_by_user_id
            """,
            [
                (
                    row["tutorial_group_id"],
                    row["staff_user_id"],
                    actor_user_id,
                )
                for row in valid_rows
            ],
        )
        record_audit_event(
            conn,
            "tutorial_group.staff_mapping_imported",
            "unit_offering",
            unit_offering_id,
            actor_user_id=actor_user_id,
            metadata={
                "source_file_name": upload.original_file_name,
                "mapping_count": len(valid_rows),
                "added_count": added_count,
                "unit_staff_added_count": len(unit_staff_ids_to_add),
            },
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return {
        "tutorial_groups": get_tutorial_group_overview(
            conn,
            actor_user_id,
            unit_offering_id,
        ),
        "staff_import": {
            "source_file_name": upload.original_file_name,
            "mapping_count": len(valid_rows),
            "added_count": added_count,
            "unit_staff_added_count": len(unit_staff_ids_to_add),
            "unchanged_count": len(valid_rows) - added_count,
        },
    }


def create_tutorial_group_import(
    conn: sqlite3.Connection,
    actor_user_id: int,
    unit_offering_id: int,
    upload: StoredUpload,
    activity_type: str = "Applied",
) -> dict:
    _unit_context(conn, actor_user_id, unit_offering_id)
    requested_activity = str(activity_type or "Applied").strip()
    if upload.extension in {".xls", ".xlsx"} and requested_activity.casefold() != "applied":
        raise ApiError(
            "tutorial_activity_unsupported",
            "Only Applied Allocate+ worksheets are currently supported.",
            422,
        )
    selected_activity = "Applied"
    source_content_hash = upload.content_hash
    if upload.extension in {".xls", ".xlsx"}:
        source_content_hash = hashlib.sha256(
            f"{upload.content_hash}:{selected_activity.casefold()}".encode("utf-8")
        ).hexdigest()
    existing = conn.execute(
        """
        SELECT tutorial_group_import_id
        FROM tutorial_group_imports
        WHERE unit_offering_id = ? AND source_content_hash = ?
        """,
        (unit_offering_id, source_content_hash),
    ).fetchone()
    if existing is not None:
        remove_stored_upload(upload.storage_path)
        return get_tutorial_group_import(
            conn,
            actor_user_id,
            int(existing["tutorial_group_import_id"]),
        )

    headers, raw_rows = _read_group_source(upload, selected_activity)
    by_key = {_header_key(header): header for header in headers}
    student_column = next(
        (
            by_key[key]
            for key in (
                "student_id",
                "student_code",
                "student_identifier",
                "student_number",
                "username",
            )
            if key in by_key
        ),
        None,
    )
    group_column = next(
        (
            by_key[key]
            for key in ("tutorial_group", "tutorial", "group", "class")
            if key in by_key
        ),
        None,
    )
    if student_column is None or group_column is None:
        raise ApiError(
            "tutorial_csv_columns",
            "CSV columns must include student_id and tutorial_group.",
            422,
        )
    registration_columns = (
        _embedded_registration_columns(by_key)
        if upload.extension in {".xls", ".xlsx"}
        else None
    )
    if upload.extension in {".xls", ".xlsx"} and registration_columns is None:
        raise ApiError(
            "tutorial_excel_student_identity_columns",
            (
                "Applied Allocate+ worksheets must include student_code, "
                "last_name, preferred_name and email_address."
            ),
            422,
        )
    normalized = [
        {
            "source_row_number": int(row.get("_source_row_number") or number),
            "raw": row,
            "student_identifier": str(row.get(student_column) or "").strip(),
            "group_code": str(row.get(group_column) or "").strip(),
            "registration": _registration_values(row, registration_columns),
        }
        for number, row in enumerate(raw_rows, start=2)
    ]
    identifier_counts = Counter(
        row["student_identifier"].casefold()
        for row in normalized
        if row["student_identifier"]
    )
    email_counts = Counter(
        str(row["registration"].get("email") or "").casefold()
        for row in normalized
        if row["registration"].get("email")
    )
    existing_students = {
        str(row["institution_student_identifier"]).casefold(): row
        for row in conn.execute("SELECT * FROM students")
    }
    student_email_owners = {
        str(row["institution_email"]).casefold(): int(row["student_id"])
        for row in conn.execute(
            """
            SELECT student_id, institution_email
            FROM students
            WHERE institution_email IS NOT NULL
              AND trim(institution_email) != ''
            """
        )
    }
    user_email_owners = {
        str(row["email"]).casefold(): int(row["user_id"])
        for row in conn.execute("SELECT user_id, email FROM users")
    }
    enrolments = {
        int(row["student_id"]): row
        for row in conn.execute(
            """
            SELECT * FROM student_enrolments
            WHERE unit_offering_id = ?
            """,
            (unit_offering_id,),
        )
    }
    memberships = {
        int(row["student_id"]): row
        for row in conn.execute(
            """
            SELECT membership.student_id, membership.tutorial_group_id,
                   group_row.group_code
            FROM student_tutorial_memberships AS membership
            JOIN tutorial_groups AS group_row
              ON group_row.tutorial_group_id = membership.tutorial_group_id
            WHERE membership.unit_offering_id = ?
              AND membership.active = 1
            """,
            (unit_offering_id,),
        )
    }
    existing_groups = {
        str(row["group_code"]).casefold(): int(row["tutorial_group_id"])
        for row in conn.execute(
            """
            SELECT tutorial_group_id, group_code
            FROM tutorial_groups
            WHERE unit_offering_id = ?
            """,
            (unit_offering_id,),
        )
    }
    prepared = []
    for row in normalized:
        key = row["student_identifier"].casefold()
        student = existing_students.get(key)
        registration = dict(row["registration"])
        error = None
        if not row["student_identifier"]:
            error = "Student ID is required."
        elif identifier_counts[key] > 1:
            error = "Student ID is duplicated in the CSV."
        elif not row["group_code"]:
            error = "Tutorial Group is required."
        if registration["available"]:
            email = str(registration.get("email") or "")
            full_name = str(registration.get("full_name") or "")
            current_student_id = int(student["student_id"]) if student else None
            linked_user_id = (
                int(student["user_id"])
                if student is not None and student["user_id"] is not None
                else None
            )
            if error is None and not full_name:
                error = "Preferred name or last name is required."
            elif error is None and (not email or "@" not in email):
                error = "A valid student email address is required."
            elif error is None and email_counts[email] > 1:
                error = "Student email is duplicated in the Applied workbook."
            elif error is None and email in student_email_owners and (
                current_student_id is None
                or student_email_owners[email] != current_student_id
            ):
                error = "Student email belongs to another student."
            elif error is None and email in user_email_owners and (
                linked_user_id is None or user_email_owners[email] != linked_user_id
            ):
                error = "Student email belongs to another account."
            if student is None:
                registration["student_action"] = "create"
                registration["enrolment_action"] = "enrol"
            else:
                registration["student_action"] = (
                    "update"
                    if str(student["full_name"] or "") != full_name
                    or str(student["institution_email"] or "").casefold() != email
                    else "unchanged"
                )
                enrolment = enrolments.get(int(student["student_id"]))
                if enrolment is None:
                    registration["enrolment_action"] = "enrol"
                elif enrolment["status"] != "active":
                    registration["enrolment_action"] = "reactivate"
                else:
                    registration["enrolment_action"] = "unchanged"
        elif student is None or (
            int(student["student_id"]) not in enrolments
            or enrolments[int(student["student_id"])]["status"] != "active"
        ):
            error = "Student is not actively enrolled in this Unit."
        if error:
            action = "invalid"
            student_id = int(student["student_id"]) if student is not None else None
            previous_group_id = None
        else:
            student_id = int(student["student_id"]) if student is not None else None
            current = memberships.get(student_id) if student_id is not None else None
            previous_group_id = (
                int(current["tutorial_group_id"]) if current is not None else None
            )
            if current is None:
                action = "assign"
            elif str(current["group_code"]).casefold() == row["group_code"].casefold():
                action = "unchanged"
            else:
                action = "move"
        prepared.append(
            {
                **row,
                "student_id": student_id,
                "previous_group_id": previous_group_id,
                "tutorial_group_id": existing_groups.get(row["group_code"].casefold()),
                "action": action,
                "error": error,
                "registration": registration,
            }
        )
    counts = Counter(row["action"] for row in prepared)
    conn.execute("BEGIN")
    try:
        import_id = int(
            conn.execute(
                """
                INSERT INTO tutorial_group_imports
                    (unit_offering_id, uploaded_by_user_id,
                     source_file_name, source_file_path, source_content_hash,
                     total_row_count, valid_row_count, invalid_row_count,
                     assigned_count, moved_count, unchanged_count)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    unit_offering_id,
                    actor_user_id,
                    upload.original_file_name,
                    str(upload.storage_path),
                    source_content_hash,
                    len(prepared),
                    len(prepared) - counts["invalid"],
                    counts["invalid"],
                    counts["assign"],
                    counts["move"],
                    counts["unchanged"],
                ),
            ).lastrowid
        )
        conn.executemany(
            """
            INSERT INTO tutorial_group_import_rows
                (tutorial_group_import_id, source_row_number, raw_data_json,
                 institution_student_identifier, group_code, student_id,
                 previous_tutorial_group_id, tutorial_group_id,
                 action, validation_error)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    import_id,
                    row["source_row_number"],
                    json.dumps(
                        {
                            **row["raw"],
                            "_registration": row["registration"],
                        },
                        ensure_ascii=False,
                    ),
                    row["student_identifier"] or None,
                    row["group_code"] or None,
                    row["student_id"],
                    row["previous_group_id"],
                    row["tutorial_group_id"],
                    row["action"],
                    row["error"],
                )
                for row in prepared
            ],
        )
        record_audit_event(
            conn,
            "tutorial_group_import.previewed",
            "tutorial_group_import",
            import_id,
            actor_user_id=actor_user_id,
            metadata={
                "unit_offering_id": unit_offering_id,
                "invalid_count": counts["invalid"],
                "moved_count": counts["move"],
            },
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return get_tutorial_group_import(conn, actor_user_id, import_id)


def get_tutorial_group_import(
    conn: sqlite3.Connection,
    actor_user_id: int,
    tutorial_group_import_id: int,
) -> dict:
    import_row = conn.execute(
        """
        SELECT * FROM tutorial_group_imports
        WHERE tutorial_group_import_id = ?
        """,
        (tutorial_group_import_id,),
    ).fetchone()
    if import_row is None:
        raise ApiError(
            "tutorial_group_import_not_found",
            "Tutorial Group import not found.",
            404,
        )
    _unit_context(conn, actor_user_id, int(import_row["unit_offering_id"]))
    all_rows = [
        dict(row)
        for row in conn.execute(
            """
            SELECT import_row.*, previous_group.group_code AS previous_group_code
            FROM tutorial_group_import_rows AS import_row
            LEFT JOIN tutorial_groups AS previous_group
              ON previous_group.tutorial_group_id =
                 import_row.previous_tutorial_group_id
            WHERE import_row.tutorial_group_import_id = ?
            ORDER BY CASE import_row.action
                       WHEN 'invalid' THEN 0 WHEN 'move' THEN 1 ELSE 2 END,
                      import_row.source_row_number
            """,
            (tutorial_group_import_id,),
        )
    ]
    registration_counts: Counter[str] = Counter()
    source_identifiers = set()
    uses_embedded_registration = False
    for row in all_rows:
        registration = _stored_registration(row.get("raw_data_json"))
        row["full_name"] = registration.get("full_name")
        row["institution_email"] = registration.get("email")
        row["student_action"] = registration.get("student_action")
        row["enrolment_action"] = registration.get("enrolment_action")
        if registration.get("available"):
            uses_embedded_registration = True
            if row["action"] != "invalid":
                registration_counts[
                    f"student:{registration.get('student_action')}"
                ] += 1
                registration_counts[
                    f"enrolment:{registration.get('enrolment_action')}"
                ] += 1
        identifier = str(row.get("institution_student_identifier") or "").casefold()
        if identifier:
            source_identifiers.add(identifier)

    not_in_applied = []
    if uses_embedded_registration:
        not_in_applied = [
            dict(row)
            for row in conn.execute(
                """
                SELECT student.institution_student_identifier,
                       student.full_name, student.institution_email
                FROM student_enrolments AS enrolment
                JOIN students AS student ON student.student_id = enrolment.student_id
                WHERE enrolment.unit_offering_id = ?
                  AND enrolment.status = 'active'
                ORDER BY student.institution_student_identifier
                """,
                (import_row["unit_offering_id"],),
            )
            if str(row["institution_student_identifier"]).casefold()
            not in source_identifiers
        ]
    return {
        "tutorial_group_import": dict(import_row),
        "rows": all_rows[:250],
        "student_registration": {
            "available": uses_embedded_registration,
            "created_count": registration_counts["student:create"],
            "updated_count": registration_counts["student:update"],
            "enrolled_count": registration_counts["enrolment:enrol"],
            "reactivated_count": registration_counts["enrolment:reactivate"],
            "not_in_applied_count": len(not_in_applied),
            "not_in_applied_students": not_in_applied[:100],
            "missing_students_unchanged": True,
        },
    }


def apply_tutorial_group_import(
    conn: sqlite3.Connection,
    actor_user_id: int,
    tutorial_group_import_id: int,
) -> dict:
    import_row = conn.execute(
        """
        SELECT * FROM tutorial_group_imports
        WHERE tutorial_group_import_id = ?
        """,
        (tutorial_group_import_id,),
    ).fetchone()
    if import_row is None:
        raise ApiError(
            "tutorial_group_import_not_found",
            "Tutorial Group import not found.",
            404,
        )
    unit_offering_id = int(import_row["unit_offering_id"])
    _unit_context(conn, actor_user_id, unit_offering_id)
    if import_row["status"] == "applied":
        return get_tutorial_group_overview(conn, actor_user_id, unit_offering_id)
    if import_row["status"] != "previewed":
        raise ApiError(
            "tutorial_group_import_state",
            "This Tutorial Group import cannot be applied.",
            409,
        )
    if int(import_row["invalid_row_count"]):
        raise ApiError(
            "tutorial_group_import_invalid",
            "Fix every invalid CSV row before applying the import.",
            409,
            {"invalid_row_count": int(import_row["invalid_row_count"])},
        )
    rows = conn.execute(
        """
        SELECT * FROM tutorial_group_import_rows
        WHERE tutorial_group_import_id = ?
        ORDER BY source_row_number
        """,
        (tutorial_group_import_id,),
    ).fetchall()
    group_names: dict[str, str] = {}
    registrations: dict[int, dict] = {}
    registration_counts: Counter[str] = Counter()
    for row in rows:
        raw = json.loads(row["raw_data_json"] or "{}")
        group_name = str(raw.get("_group_name") or "").strip()
        if group_name:
            group_names.setdefault(str(row["group_code"]).casefold(), group_name)
        registration = _stored_registration(row["raw_data_json"])
        registrations[int(row["tutorial_group_import_row_id"])] = registration
        if registration.get("available"):
            registration_counts[
                f"student:{registration.get('student_action')}"
            ] += 1
            registration_counts[
                f"enrolment:{registration.get('enrolment_action')}"
            ] += 1
    conn.execute("BEGIN")
    try:
        group_ids: dict[str, int] = {}
        for group_code in sorted(
            {str(row["group_code"]) for row in rows},
            key=str.casefold,
        ):
            existing_group = conn.execute(
                """
                SELECT tutorial_group_id
                FROM tutorial_groups
                WHERE unit_offering_id = ? AND group_code = ?
                """,
                (unit_offering_id, group_code),
            ).fetchone()
            if existing_group is None:
                group_id = int(
                    conn.execute(
                        """
                        INSERT INTO tutorial_groups
                            (unit_offering_id, group_code, group_name,
                             source, created_by_user_id)
                        VALUES (?, ?, ?, 'csv', ?)
                        """,
                        (
                            unit_offering_id,
                            group_code,
                            group_names.get(group_code.casefold(), group_code),
                            actor_user_id,
                        ),
                    ).lastrowid
                )
            else:
                group_id = int(existing_group["tutorial_group_id"])
                conn.execute(
                    """
                    UPDATE tutorial_groups
                    SET active = 1, updated_at = CURRENT_TIMESTAMP
                    WHERE tutorial_group_id = ?
                    """,
                    (group_id,),
                )
            group_ids[group_code.casefold()] = group_id
        for row in rows:
            registration = registrations[int(row["tutorial_group_import_row_id"])]
            if registration.get("available"):
                student_id = _apply_embedded_student_registration(
                    conn,
                    unit_offering_id,
                    tutorial_group_import_id,
                    str(row["institution_student_identifier"] or "").strip(),
                    registration,
                )
            elif row["student_id"] is not None:
                student_id = int(row["student_id"])
            else:
                raise ApiError(
                    "tutorial_student_missing",
                    "An imported student is no longer available.",
                    409,
                )
            group_id = group_ids[str(row["group_code"]).casefold()]
            current = conn.execute(
                """
                SELECT student_tutorial_membership_id, tutorial_group_id
                FROM student_tutorial_memberships
                WHERE unit_offering_id = ? AND student_id = ? AND active = 1
                """,
                (unit_offering_id, student_id),
            ).fetchone()
            if current is None or int(current["tutorial_group_id"]) != group_id:
                if current is not None:
                    conn.execute(
                        """
                        UPDATE student_tutorial_memberships
                        SET active = 0, ended_at = CURRENT_TIMESTAMP
                        WHERE student_tutorial_membership_id = ?
                        """,
                        (current["student_tutorial_membership_id"],),
                    )
                conn.execute(
                    """
                    INSERT INTO student_tutorial_memberships
                        (unit_offering_id, student_id, tutorial_group_id,
                         source, tutorial_group_import_id)
                    VALUES (?, ?, ?, 'csv', ?)
                    """,
                    (
                        unit_offering_id,
                        student_id,
                        group_id,
                        tutorial_group_import_id,
                    ),
                )
            conn.execute(
                """
                UPDATE tutorial_group_import_rows
                SET student_id = ?, tutorial_group_id = ?,
                    applied_at = CURRENT_TIMESTAMP
                WHERE tutorial_group_import_row_id = ?
                """,
                (student_id, group_id, row["tutorial_group_import_row_id"]),
            )
        conn.execute(
            """
            UPDATE tutorial_group_imports
            SET status = 'applied', applied_at = CURRENT_TIMESTAMP
            WHERE tutorial_group_import_id = ?
            """,
            (tutorial_group_import_id,),
        )
        record_audit_event(
            conn,
            "tutorial_group_import.applied",
            "tutorial_group_import",
            tutorial_group_import_id,
            actor_user_id=actor_user_id,
            metadata={
                "unit_offering_id": unit_offering_id,
                "assigned_count": int(import_row["assigned_count"]),
                "moved_count": int(import_row["moved_count"]),
                "created_student_count": registration_counts["student:create"],
                "updated_student_count": registration_counts["student:update"],
                "enrolled_student_count": registration_counts["enrolment:enrol"],
                "reactivated_student_count": registration_counts[
                    "enrolment:reactivate"
                ],
                "missing_students_unchanged": True,
                "existing_allocations_preserved": True,
            },
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return get_tutorial_group_overview(conn, actor_user_id, unit_offering_id)
